import os
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

def select_xlsx_files(extension='.xlsx'):
    """엑셀 파일 선택 함수"""
    file_list = [file for file in os.listdir() if file.endswith(extension)]
    if not file_list:
        print("no excel file")
        return None
    
    for i, file in enumerate(file_list, start=1):
        print(f"{i}. {file}")
    
    while True:
        choice = input("input file number (exit: 0) ")
        if choice.isdigit():
            choice = int(choice)
            if choice == 0:
                print('exit the program')
                return None
            elif 1 <= choice <= len(file_list):
                return file_list[choice -1]
        print('Invalid number. try again.')

def update_version(filename: str, final_version: bool = False) -> str:
    """파일 버전 업데이트 함수"""
    base_name, ext = filename.rsplit('.', 1)

    match = re.search(r'_v(\d+)$', base_name)
    final_match = re.search(r'_vf$', base_name)

    if final_match:
        return filename
    
    if final_version:
        if match:
            new_base_name = re.sub(r'_v\d+$', '_vf', base_name)
        else:
            new_base_name = f"{base_name}_vf"
    else:
        if match:
            version = int(match.group(1))
            new_version = version + 1
            new_base_name = re.sub(r'_v\d+$', f'_v{new_version}', base_name)
        else:
            new_base_name = f"{base_name}_v1"
    
    return f"{new_base_name}.{ext}"

def save_to_excel(df, type, file_name):
    """엑셀 파일 저장 및 스타일링"""
    wb = load_workbook(file_name)
    sheet = wb[type]

    sheet.insert_rows(1)
    sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'
    sheet['A1'].font = Font(bold=True)

    for col in range(1, 8):
        cell = sheet.cell(row=2, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    if type != '이력없음_미사용정책':
        for col in range(8, 24):
            cell = sheet.cell(row=2, column=col)
            cell.fill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
    
    wb.save(file_name)

def remove_extension(filename):
    """파일 확장자 제거"""
    return os.path.splitext(filename)[0]

def convert_to_date(date_str):
    """날짜 문자열을 날짜 형식으로 변환"""
    try:
        date_obj = datetime.strptime(date_str, '%Y%m%d')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError:
        return date_str

def find_auto_extension_id():
    """자동 연장 ID 찾기"""
    print('가공된 신청정보 파일을 선택')
    selected_file = select_xlsx_files()
    df = pd.read_excel(selected_file)
    filtered_df = df[df['REQUEST_STATUS'].isin([98, 99])]['REQUEST_ID'].drop_duplicates()
    return filtered_df 